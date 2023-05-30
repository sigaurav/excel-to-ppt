from django.core.files.storage import FileSystemStorage
import os
import win32com.client
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from .process_file import Generate_ppt



def upload_file(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        try:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            sheet_list = wb.sheetnames

            # Delete all files within the inputFiles directory
            input_files_directory = os.path.join(settings.MEDIA_ROOT, 'inputFiles')

            return JsonResponse({'sheet_list': sheet_list})
        except InvalidFileException:
            return JsonResponse({'error': 'Invalid Excel file. Please upload a valid Excel file.'}, status=400)

    return render(request, 'upload_file.html')


def generate_ppt(request):
    if request.method == 'POST':
        selected_sheets = request.POST.getlist('sheet_list_html')
        filename = request.POST.get('filename')
        print(filename)
        file_path = os.path.join('excelapp/inputFiles/', filename)
        print(file_path)

        generate_ppt = Generate_ppt(file_path, selected_sheets)
        output_path = os.path.join(settings.MEDIA_ROOT, 'outputPath')
        # output_path = generate_ppt.generate_ppt(file_path, selected_sheets)

        return JsonResponse({'success': True, 'output_path': output_path})

    return render(request, 'upload_file.html')



def download_ppt(request):
    if request.method == 'GET':
        output_path = request.GET.get('output_path', '')

        if output_path:
            with open(output_path, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation')
                response['Content-Disposition'] = 'attachment; filename=output.pptx'
                return response

    return render(request, 'upload_file.html')
